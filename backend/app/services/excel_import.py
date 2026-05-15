"""Parse product rows and optional trailer/scenario meta from .xlsx / .csv uploads."""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook, load_workbook

from app.data.scenarios import STANDARD_TRAILER
from app.models.schemas import AxleLoadLimits, Product, Trailer
from app.services.import_meta import ImportMeta

PRODUCT_SHEETS = ("products", "produkty", "ladunek", "ładunek", "cargo", "items")
TRAILER_SHEETS = ("trailer", "naczepa", "vehicle", "pojazd")


def _bool_cell(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "tak")


def _float_cell(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _int_cell(v: Any, default: int = 1) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _norm_key(k: str) -> str:
    return "".join(ch for ch in k.lower().strip() if ch.isalnum())


def _parse_product_row(raw: tuple, idx: dict[str, int], row_num: int) -> Product | None:
    if raw is None or all(c is None or str(c).strip() == "" for c in raw):
        return None

    def g(*names: str, default=None):
        for n in names:
            nk = _norm_key(n)
            if nk in idx:
                return raw[idx[nk]]
        return default

    name = g("productname", "name", "nazwa", default=None)
    if name is None or str(name).strip() == "":
        return None
    pid = str(g("productid", "sku", "kod", default=f"IMP-{row_num}"))
    pkg = str(g("packagingkind", "packaging", default="rigid") or "rigid").strip().lower()
    if pkg not in ("rigid", "compressible", "max_packed"):
        pkg = "rigid"
    return Product(
        product_id=pid,
        name=str(name).strip(),
        length_mm=_float_cell(g("lengthmm", "length", "dlugosc"), 400),
        width_mm=_float_cell(g("widthmm", "width", "szerokosc"), 300),
        height_mm=_float_cell(g("heightmm", "height", "wysokosc"), 250),
        weight_kg=_float_cell(g("weightkg", "weight", "waga"), 10),
        quantity=max(1, _int_cell(g("quantity", "qty", "ilosc", "szt"), 1)),
        fragile=_bool_cell(g("fragile", "fragility", "kruche")),
        compressible=_bool_cell(g("compressible", "compress", "sciskalne")),
        max_stack_weight_kg=_float_cell(g("maxstackweightkg", "maxstack", "maxstos"), 500) or None,
        can_rotate=not _bool_cell(g("norotate", "fixedorientation", "bezobrotu")),
        stacking_group=str(g("stackinggroup", "group", "grupastosowania") or "IMPORT"),
        packaging_kind=pkg,  # type: ignore[arg-type]
        internal_void_ratio=min(0.85, max(0.0, _float_cell(g("internalvoidratio", "voidratio", "luz"), 0))),
    )


def _products_from_rows(rows: list[tuple]) -> list[Product]:
    if not rows:
        return []
    header = [_norm_key(str(c or "")) for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    products: list[Product] = []
    for i, raw in enumerate(rows[1:], start=2):
        p = _parse_product_row(raw, idx, i)
        if p:
            products.append(p)
    return products


def _kv_from_rows(rows: list[tuple]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for raw in rows:
        if not raw or len(raw) < 2:
            continue
        k = _norm_key(str(raw[0] or ""))
        if not k or k in ("pole", "field", "klucz", "key"):
            continue
        out[k] = raw[1]
    return out


def _meta_from_kv(kv: dict[str, Any]) -> ImportMeta:
    return ImportMeta(
        scenario_id=str(kv["scenarioid"]).strip() if kv.get("scenarioid") not in (None, "") else None,
        title=str(kv["scenariotitle"]).strip() if kv.get("scenariotitle") not in (None, "") else None,
        description=str(kv["scenariodescription"]).strip()
        if kv.get("scenariodescription") not in (None, "")
        else None,
    )


def _trailer_from_kv(kv: dict[str, Any], base: Trailer) -> Trailer:
    front = _float_cell(kv.get("frontaxlelimitkg") or kv.get("frontaxlekg"), base.axle_load_limits.front_kg)
    rear = _float_cell(kv.get("rearaxlelimitkg") or kv.get("rearaxlekg"), base.axle_load_limits.rear_kg)
    return Trailer(
        trailer_id=str(kv.get("trailerid") or base.trailer_id),
        name=str(kv.get("trailername") or kv.get("name") or base.name),
        length_mm=_float_cell(kv.get("lengthmm"), base.length_mm),
        width_mm=_float_cell(kv.get("widthmm"), base.width_mm),
        height_mm=_float_cell(kv.get("heightmm"), base.height_mm),
        max_weight_kg=_float_cell(kv.get("maxweightkg"), base.max_weight_kg),
        max_stack_height_mm=_float_cell(kv.get("maxstackheightmm"), base.max_stack_height_mm),
        axle_load_limits=AxleLoadLimits(front_kg=front, rear_kg=rear),
        wheelbase_mm=_float_cell(kv.get("wheelbasemm"), base.wheelbase_mm),
        track_width_mm=_float_cell(kv.get("trackwidthmm"), base.track_width_mm),
        deck_height_mm=_float_cell(kv.get("deckheightmm"), base.deck_height_mm),
        max_lateral_accel_g=_float_cell(kv.get("maxlateralaccelg"), base.max_lateral_accel_g),
        max_brake_accel_g=_float_cell(kv.get("maxbrakeaccelg"), base.max_brake_accel_g),
    )


def _find_sheet_name(names: list[str], candidates: tuple[str, ...]) -> str | None:
    norm = {_norm_key(n): n for n in names}
    for c in candidates:
        if c in norm:
            return norm[c]
    return None


def parse_excel(content: bytes) -> tuple[Trailer, list[Product], ImportMeta]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    meta = ImportMeta()
    trailer = STANDARD_TRAILER

    trailer_sheet = _find_sheet_name(sheet_names, TRAILER_SHEETS)
    if trailer_sheet:
        ws_t = wb[trailer_sheet]
        kv = _kv_from_rows(list(ws_t.iter_rows(values_only=True)))
        meta = _meta_from_kv(kv)
        trailer = _trailer_from_kv(kv, trailer)

    product_sheet = _find_sheet_name(sheet_names, PRODUCT_SHEETS)
    if not product_sheet:
        for n in sheet_names:
            if trailer_sheet and n == trailer_sheet:
                continue
            product_sheet = n
            break
    if not product_sheet:
        wb.close()
        return trailer, [], meta

    ws_p = wb[product_sheet]
    products = _products_from_rows(list(ws_p.iter_rows(values_only=True)))
    wb.close()
    return trailer, products, meta


def parse_csv_text(text: str) -> tuple[Trailer, list[Product], ImportMeta]:
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return STANDARD_TRAILER, [], ImportMeta()
    fmap = {_norm_key(h): h for h in reader.fieldnames}

    def col(*candidates: str) -> str | None:
        for c in candidates:
            nk = _norm_key(c)
            if nk in fmap:
                return fmap[nk]
        return None

    products: list[Product] = []
    for i, row in enumerate(reader, start=2):

        def gv(*names: str, default=None):
            k = col(*names)
            if not k:
                return default
            return row.get(k, default)

        name = gv("ProductName", "name", "Nazwa", default=None)
        if not name or str(name).strip() == "":
            continue
        pid = str(gv("productid", "sku", default=f"IMP-{i}"))
        pkg = str(gv("packagingkind", default="rigid") or "rigid").strip().lower()
        if pkg not in ("rigid", "compressible", "max_packed"):
            pkg = "rigid"
        products.append(
            Product(
                product_id=pid,
                name=str(name).strip(),
                length_mm=_float_cell(gv("LengthMm", "length"), 400),
                width_mm=_float_cell(gv("WidthMm", "width"), 300),
                height_mm=_float_cell(gv("HeightMm", "height"), 250),
                weight_kg=_float_cell(gv("WeightKg", "weight"), 10),
                quantity=max(1, _int_cell(gv("Quantity", "qty", "Ilosc"), 1)),
                fragile=_bool_cell(gv("Fragile")),
                compressible=_bool_cell(gv("Compressible")),
                max_stack_weight_kg=_float_cell(gv("MaxStackWeightKg", "maxstack"), 500) or None,
                can_rotate=not _bool_cell(gv("NoRotate")),
                stacking_group=str(gv("StackingGroup", "group") or "IMPORT"),
                packaging_kind=pkg,  # type: ignore[arg-type]
                internal_void_ratio=min(0.85, max(0.0, _float_cell(gv("InternalVoidRatio"), 0))),
            )
        )
    return STANDARD_TRAILER, products, ImportMeta()


def build_scenario_template_xlsx() -> bytes:
    """Generuje plik szablonu .xlsx (Products + Trailer)."""
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "Products"
    ws_p.append(
        [
            "ProductId",
            "ProductName",
            "LengthMm",
            "WidthMm",
            "HeightMm",
            "WeightKg",
            "Quantity",
            "Fragile",
            "Compressible",
            "MaxStackWeightKg",
            "StackingGroup",
            "NoRotate",
            "PackagingKind",
            "InternalVoidRatio",
        ]
    )
    ws_p.append(["BOX-A", "Karton A (przykład)", 1200, 800, 900, 180, 6, "NIE", "NIE", 2000, "GENERAL", "NIE", "rigid", 0])
    ws_p.append(["BOX-B", "Karton B (przykład)", 1000, 1000, 800, 220, 2, "NIE", "TAK", 1500, "GENERAL", "NIE", "compressible", 0.1])
    ws_p.append(["", "", "", "", "", "", "", "", "", "", "", "", "", ""])

    ws_t = wb.create_sheet("Trailer")
    ws_t.append(["Pole", "Wartość"])
    trailer_rows = [
        ("ScenarioId", "MOJ_SCENARIUSZ_01"),
        ("ScenarioTitle", "Mój załadunek — przykład"),
        ("ScenarioDescription", "Własny scenariusz z szablonu. Po imporcie użyj Przelicz / Stosy / AI."),
        ("TrailerName", "Naczepa standardowa (własna nazwa)"),
        ("LengthMm", 13600),
        ("WidthMm", 2450),
        ("HeightMm", 2700),
        ("MaxWeightKg", 24000),
        ("MaxStackHeightMm", 2700),
        ("FrontAxleLimitKg", 8000),
        ("RearAxleLimitKg", 18000),
    ]
    for row in trailer_rows:
        ws_t.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
